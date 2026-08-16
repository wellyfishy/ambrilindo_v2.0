from django.shortcuts import render, redirect, get_object_or_404 # type: ignore
from django.contrib.auth import authenticate, login, logout # type: ignore
from django.contrib import messages # type: ignore
from .models import * # type: ignore
import openpyxl # type: ignore
from django.db.models import Count, Q # type: ignore
import random # type: ignore
from channels.layers import get_channel_layer # type: ignore
from asgiref.sync import async_to_sync # type: ignore
from django.views.decorators.csrf import csrf_exempt # type: ignore
from django.http import JsonResponse, HttpResponse # type: ignore
from collections import defaultdict # type: ignore
from itertools import groupby # type: ignore
import json # type: ignore
from django.views.decorators.http import require_POST # type: ignore
from .utils import send_to_hosted 
from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font # type: ignore
from collections import Counter
import re

import io
from django.conf import settings # type: ignore
from playwright.sync_api import sync_playwright # type: ignore

from pypdf import PdfWriter, PdfReader # type: ignore
from django.urls import reverse # type: ignore

import sys
import asyncio
import concurrent.futures


def auth(request):
    events = Event.objects.all().order_by('-pk')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        event_pk = request.POST.get('event_pk')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            try:
                event = Event.objects.get(pk=event_pk)
                admin_tatami = AdminTatami.objects.filter(user=user, event=event).first()
                admin = Admin.objects.filter(user=user).first()
                if admin_tatami:
                    login(request, user)
                    return redirect('admin-dashboard', event_pk=event_pk)
                elif admin:
                    login(request, user)
                    return redirect('admin-dashboard', event_pk=event_pk)
                else:
                    jury = Jury.objects.filter(event=event, user=user).first()
                    if jury:
                        login(request, user)
                        return redirect('jury-panel', tatami_pk=jury.tatami.pk)
                    else:
                        messages.error(request, "Anda tidak terdaftar sebagai Admin atau Juri untuk event ini.")
                        return redirect('auth')

            except Event.DoesNotExist:
                messages.error(request, "Event tidak ditemukan.")
                return redirect('auth')
        else:
            if 'c' in username:
                new_username = username.replace('c', '')
                coach_supervisor = Tatami.objects.filter(event__pk=new_username, tatami_number=password).first()
                if coach_supervisor:
                    return redirect('coach-supervisor', tatami_pk=coach_supervisor.pk)
            else:
                adm_control = Tatami.objects.filter(event__pk=username, tatami_number=password).first()
                if adm_control:
                    return redirect('admin-control', tatami_pk=adm_control.pk)
            messages.error(request, "Username atau password salah!")
            return redirect('auth')
        
    context = {
        'events': events
    }
        
    return render(request, 'auth/auth.html', context)

def admin_control(request, tatami_pk):
    tatami = Tatami.objects.get(pk=tatami_pk)
    event = tatami.event
    context = {
        'tatami': tatami,
        'on': 'control',
        'event': event,
    }
    return render(request, 'jury/admin-control.html', context)

def jury_panel(request, tatami_pk):
    jury = Jury.objects.get(user=request.user)
    detail_bagan = jury.tatami.detail_bagan
    tatami = Tatami.objects.get(pk=tatami_pk)

    context = {
        'jury': jury,
        'detail_bagan': detail_bagan,
        'tatami': tatami,
    }
    return render(request, 'jury/jury-panel.html', context)

def coach_supervisor(request, tatami_pk):
    tatami = Tatami.objects.get(pk=tatami_pk)

    context = {
        'tatami': tatami,
    }
    return render(request, 'jury/coach-supervisor.html', context)

@csrf_exempt
def message_retriever_jury(request, tatami_pk):
    if request.method == 'POST':
        action = request.POST.get('action')
        details = request.POST.get('details')

        group_name = f"juryroom_{tatami_pk}"
        channel_layer = get_channel_layer()

        async_to_sync(channel_layer.group_send)(
            group_name,
            {
                "type": "broadcast_command",
                "message": action,
                "details": details,
            }
        )

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'error': 'Invalid method'}, status=405)

@csrf_exempt
def message_retriever_admin(request, tatami_pk):
    if request.method == 'POST':
        action = request.POST.get('action')
        details = request.POST.get('details')

        group_name = f"control_{tatami_pk}"
        channel_layer = get_channel_layer()

        async_to_sync(channel_layer.group_send)(
            group_name,
            {
                "type": "broadcast_command",
                "message": action,
                "details": details,
            }
        )

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'error': 'Invalid method'}, status=405)

@csrf_exempt
def message_retriever_control(request, tatami_pk):
    if request.method == 'POST':
        action = request.POST.get('action')
        details = request.POST.get('details')

        group_name = f"admin_control_{tatami_pk}"
        channel_layer = get_channel_layer()

        async_to_sync(channel_layer.group_send)(
            group_name,
            {
                "type": "broadcast_command",
                "message": action,
                "details": details,
            }
        )

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'error': 'Invalid method'}, status=405)

@csrf_exempt
def message_retriever_coach_supervisor(request, tatami_pk):
    if request.method == 'POST':
        action = request.POST.get('action')
        details = request.POST.get('details')

        group_name = f"coachroom_{tatami_pk}"
        channel_layer = get_channel_layer()

        async_to_sync(channel_layer.group_send)(
            group_name,
            {
                "type": "broadcast_command",
                "message": action,
                "details": details,
            }
        )

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'error': 'Invalid method'}, status=405)

def get_current_atlets(request, tatami_pk):
    tatami = Tatami.objects.get(pk=tatami_pk)
    detail_bagan = tatami.detail_bagan
    data = {
        'tatami': tatami.pk,
        'detail_bagan': detail_bagan.pk,
        'atlet1_nama': detail_bagan.atlet1.nama_atlet,
        'atlet1_perguruan': detail_bagan.atlet1.perguruan,
        'atlet1_utusan': detail_bagan.atlet1.utusan,
        'atlet1_vr': detail_bagan.vr1,
        'atlet2_nama': detail_bagan.atlet2.nama_atlet,
        'atlet2_perguruan': detail_bagan.atlet2.perguruan,
        'atlet2_utusan': detail_bagan.atlet2.utusan,
        'atlet2_vr': detail_bagan.vr2,
    }

    return JsonResponse(data)

def logoutfunc(request):
    logout(request)
    return redirect('auth')

def admin_dashboard(request, event_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    nomor_tandings = NomorTanding.objects.filter(event=event)
    for nt in nomor_tandings:
        nt.jumlat_atlet = Atlet.objects.filter(nomor_tanding=nt).count()
    
    bagans = Bagan.objects.filter(event=event).order_by('-kode')
    
    # count = 0
    # for bagan in bagans:
    #     dbs = DetailBagan.objects.filter(bagan=bagan)
    #     for db in dbs:
    #         if db.atlet1:
    #             count += 1
    #         if db.atlet2:
    #             count += 1
    # print(count)

    custom_order = [1, 8, 4, 5, 2, 7, 3, 6]

    atlet_assignment1 = {
                        1: 'atlet1',
                        8: 'atlet2',
                        4: 'atlet2',
                        5: 'atlet1',
                        2: 'atlet2',
                        7: 'atlet1',
                        3: 'atlet1',
                        6: 'atlet2',
                    }

    if request.method == 'POST':
        if request.POST.get('submit_type') == 'export_bagan':
            bagan_pks = request.POST.getlist('bagan_pk')

            if 'semua' in bagan_pks:
                bagan_pks = Bagan.objects.filter(event=event).values_list('pk', flat=True)

            wb = Workbook()
            ws = wb.active
            ws.title = 'Bagan Export'

            headers = ['Kode', 'Nama Bagan', 'Nama Nomor Tanding', 'Bagan PK', 'Detail PK', 'Round', 'Urutan', 'Atlet 1', 'Perguruan 1', 'Perwakilan 1', 'Atlet 2', 'Perguruan 2', 'Perwakilan 2', 'Tipe Tanding', 'Pool', 'VR 1', 'VR 2', 'Score 1', 'Score 2', 'Status Selesai', 'Pemenang']
            ws.append(headers)

            # bold header row
            for cell in ws[1]:
                cell.font = Font(bold=True)

            for bagan_pk in bagan_pks:
                bagan = Bagan.objects.filter(pk=bagan_pk).first()
                if not bagan:
                    continue
                dbs = DetailBagan.objects.filter(bagan=bagan)
                for db in dbs:
                    nama1 = db.atlet1.nama_atlet if db.atlet1 else None
                    perguruan1 = db.atlet1.perguruan.nama_perguruan if db.atlet1 else None
                    perwakilan1 = db.atlet1.utusan.nama_utusan if db.atlet1 else None
                    nama2 = db.atlet2.nama_atlet if db.atlet2 else None
                    perguruan2 = db.atlet2.perguruan.nama_perguruan if db.atlet2 else None
                    perwakilan2 = db.atlet2.utusan.nama_utusan if db.atlet2 else None
                    ws.append([
                        bagan.kode,
                        bagan.nama_bagan,
                        bagan.nomor_tanding.nama_nomor_tanding,
                        bagan.pk,
                        db.pk,
                        db.round,
                        db.urutan,
                        nama1,
                        perguruan1,
                        perwakilan1,
                        nama2,
                        perguruan2,
                        perwakilan2,
                        bagan.tipe_tanding,
                        bagan.pool,
                        db.vr1,
                        db.vr2,
                        db.score1,
                        db.score2,
                        db.selesai,
                        db.pemenang,
                    ])

            # auto-size columns roughly
            for col_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = length + 2

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="bagan_export_{event_pk}.xlsx"'
            wb.save(response)
            return response

        elif request.POST.get('submit_type') == 'bob_bagan':
            bagan_pks = request.POST.getlist('bagan_pk')
            tipe_shuffle = request.POST.get('shuffle_type')
            nama_bob = request.POST.get('nama_bob')  # manually typed name from the form

            if not nama_bob:
                messages.error(request, "Nama kategori Best of the Best wajib diisi.")
                return redirect('admin-detail-event', event_pk=event_pk)

            bagan_list = list(Bagan.objects.filter(pk__in=bagan_pks))

            atlets_temp_all = [b.juara_1 for b in bagan_list if b.juara_1]

            if not atlets_temp_all:
                messages.error(request, "Tidak ada juara 1 pada bagan yang dipilih.")
                return redirect('admin-detail-event', event_pk=event_pk)

            if tipe_shuffle == 'perguruan':
                group_model = Perguruan
                group_field = 'perguruan'
            elif tipe_shuffle == 'kontingen':
                group_model = Utusan
                group_field = 'utusan'

            # get_or_create the new category
            nomor_tanding, _ = NomorTanding.objects.get_or_create(
                event=event,
                nama_nomor_tanding=nama_bob,
            )

            bob_atlets = []
            for atlet in atlets_temp_all:
                new_atlet = Atlet.objects.get(pk=atlet.pk)
                new_atlet.pk = None
                new_atlet.id = None
                new_atlet.nomor_tanding = nomor_tanding
                new_atlet.save()
                bob_atlets.append(new_atlet)

            atlets_temp_all = bob_atlets

            # count atletes per group, straight from the python list (no DB filter by nomor_tanding needed)
            group_counts_counter = Counter(
                getattr(atlet, f'{group_field}_id') for atlet in atlets_temp_all
            )
            group_counts = sorted(group_counts_counter.items(), key=lambda x: -x[1])

            def shuffle_same_counts(group_counts):
                result = []
                for count, g in groupby(group_counts, key=lambda x: x[1]):
                    block = list(g)
                    if len(block) > 1:
                        random.shuffle(block)
                    result.extend(block)
                return result

            group_counts = shuffle_same_counts(group_counts)
            group_counts_temp = group_counts

            def split_count_balanced(total, parts, start_index=0):
                base = total // parts
                remainder = total % parts
                splits = [base] * parts
                idx = start_index
                for _ in range(remainder):
                    splits[idx] += 1
                    idx = (idx + 1) % parts
                return splits

            if 0 < len(atlets_temp_all) < 17:
                perulangan = 1
            elif 16 < len(atlets_temp_all) < 33:
                perulangan = 2
            elif 32 < len(atlets_temp_all) < 49:
                perulangan = 3
            elif 48 < len(atlets_temp_all) < 65:
                perulangan = 4
            else:
                perulangan = 1

            if perulangan > 1 and perulangan % 2 != 0:
                perulangan += 1

            pools = [[] for _ in range(perulangan)]
            for idx, (group_id, count) in enumerate(group_counts):
                splits = split_count_balanced(count, perulangan, start_index=idx % perulangan)
                for pool_idx, val in enumerate(splits):
                    pools[pool_idx].append((group_id, val))

            group_counts_pool_a = pools[0] if perulangan >= 1 else []
            group_counts_pool_b = pools[1] if perulangan >= 2 else []
            group_counts_pool_c = pools[2] if perulangan >= 3 else []
            group_counts_pool_d = pools[3] if perulangan >= 4 else []

            def assign_atlets_to_pool(atlets, group_counts_pool, field_name):
                result = []
                remaining_counts = {gid: count for gid, count in group_counts_pool}
                for gid in remaining_counts:
                    for atlet in atlets[:]:
                        if getattr(atlet, field_name + '_id') == gid and remaining_counts[gid] > 0:
                            result.append(atlet)
                            atlets.remove(atlet)
                            remaining_counts[gid] -= 1
                return result

            atlets_temp_main = atlets_temp_all.copy()
            atlets_temp_pool_a = assign_atlets_to_pool(atlets_temp_main, group_counts_pool_a, group_field)
            atlets_temp_pool_b = assign_atlets_to_pool(atlets_temp_main, group_counts_pool_b, group_field)
            atlets_temp_pool_c = assign_atlets_to_pool(atlets_temp_main, group_counts_pool_c, group_field)
            atlets_temp_pool_d = assign_atlets_to_pool(atlets_temp_main, group_counts_pool_d, group_field)

            POOL_LETTERS = ['A', 'B', 'C', 'D']
            pool_group_counts = [group_counts_pool_a, group_counts_pool_b, group_counts_pool_c, group_counts_pool_d]
            pool_atlets = [atlets_temp_pool_a, atlets_temp_pool_b, atlets_temp_pool_c, atlets_temp_pool_d]

            final_bagan = None
            if perulangan > 1:
                final_bagan = Bagan.objects.create(
                    event=event,
                    nomor_tanding=nomor_tanding,
                    nama_bagan=f'{nama_bob} - Final',
                    pool=0,
                )
                if 'KATA' in nama_bob.upper():
                    final_bagan.tipe_tanding = '1'
                elif 'KUMITE' in nama_bob.upper():
                    final_bagan.tipe_tanding = '2'
                final_bagan.save()

                if perulangan == 2:
                    DetailBagan.objects.create(bagan=final_bagan, round=1, urutan=1)
                elif perulangan == 4:
                    DetailBagan.objects.create(bagan=final_bagan, round=1, urutan=1)
                    DetailBagan.objects.create(bagan=final_bagan, round=1, urutan=2)
                    DetailBagan.objects.create(bagan=final_bagan, round=2, urutan=1)

            pool_round5_dbs = []

            for i in range(1, perulangan + 1):
                if perulangan > 1:
                    nama_bagan = f'{nama_bob} - Pool {POOL_LETTERS[i-1]}'
                    group_counts = list(pool_group_counts[i-1])
                    atlets_temp = pool_atlets[i-1]
                else:
                    nama_bagan = nama_bob
                    group_counts = group_counts_temp
                    atlets_temp = atlets_temp_all

                bagan, round_5 = build_full_bracket(
                    event, nomor_tanding, nama_bagan, perulangan,
                    group_counts, atlets_temp, group_field, custom_order, atlet_assignment1
                )
                pool_round5_dbs.append(round_5)

            messages.success(request, f"Bagan '{nama_bob}' berhasil dibuat.")
            return redirect('admin-detail-event', event_pk=event_pk)


        elif request.POST.get('submit_type') == 'drawing_bagan':
            nomor_tanding_pks = request.POST.getlist('nomor_tanding_pk')
            tipe_shuffle = request.POST.get('shuffle_type')
            if 'semua' in nomor_tanding_pks:
                nomor_tanding_list = list(NomorTanding.objects.filter(event=event))
            else:
                nomor_tanding_list = list(NomorTanding.objects.filter(pk__in=nomor_tanding_pks))

            nomor_tanding_list.sort(key=sort_key)
            nomor_tanding_pks = [nt.pk for nt in nomor_tanding_list]
            
            if tipe_shuffle in ['perguruan', 'kontingen']:
                for nomor_tanding_pk in nomor_tanding_pks:
                    nomor_tanding = NomorTanding.objects.filter(pk=nomor_tanding_pk).first()

                    if tipe_shuffle == 'perguruan':
                        group_model = Perguruan
                        group_field = 'perguruan'
                    elif tipe_shuffle == 'kontingen':
                        group_model = Utusan
                        group_field = 'utusan'

                    group_counts = list(
                        group_model.objects.annotate(
                            num_atlet=Count('atlet', filter=Q(atlet__nomor_tanding=nomor_tanding))
                        )
                        .filter(num_atlet__gt=0)
                        .order_by('-num_atlet')
                        .values_list('id', 'num_atlet')
                    )

                    def shuffle_same_counts(group_counts):
                        result = []
                        for count, g in groupby(group_counts, key=lambda x: x[1]):
                            block = list(g)
                            if len(block) > 1:
                                random.shuffle(block)
                            result.extend(block)
                        return result

                    group_counts = shuffle_same_counts(group_counts)

                    atlets_temp_all = list(Atlet.objects.filter(nomor_tanding=nomor_tanding).filter(
                        (
                            Q(nomor_tanding__nama_nomor_tanding__icontains='kumite') &
                            Q(nomor_tanding__nama_nomor_tanding__icontains='beregu') &
                            Q(nama_atlet__icontains='team')
                        )
                        |
                        ~(
                            Q(nomor_tanding__nama_nomor_tanding__icontains='kumite') &
                            Q(nomor_tanding__nama_nomor_tanding__icontains='beregu')
                        )
                    ))
                    group_counts_temp = group_counts

                    def split_count_balanced(total, parts, start_index=0):
                        base = total // parts
                        remainder = total % parts
                        splits = [base] * parts
                        idx = start_index
                        for _ in range(remainder):
                            splits[idx] += 1
                            idx = (idx + 1) % parts
                        return splits

                    if 0 < len(atlets_temp_all) < 17:
                        perulangan = 1
                    elif 16 < len(atlets_temp_all) < 33:
                        perulangan = 2
                    elif 32 < len(atlets_temp_all) < 49:
                        perulangan = 3
                    elif 48 < len(atlets_temp_all) < 65:
                        perulangan = 4
                    else:
                        perulangan = 1

                    if perulangan > 1 and perulangan % 2 != 0:
                        perulangan += 1

                    pools = [[] for _ in range(perulangan)]
                    for idx, (group_id, count) in enumerate(group_counts):
                        splits = split_count_balanced(count, perulangan, start_index=idx % perulangan)
                        for pool_idx, val in enumerate(splits):
                            pools[pool_idx].append((group_id, val))

                    group_counts_pool_a = pools[0] if perulangan >= 1 else []
                    group_counts_pool_b = pools[1] if perulangan >= 2 else []
                    group_counts_pool_c = pools[2] if perulangan >= 3 else []
                    group_counts_pool_d = pools[3] if perulangan >= 4 else []

                    def assign_atlets_to_pool(atlets, group_counts_pool, field_name):
                        result = []
                        remaining_counts = {gid: count for gid, count in group_counts_pool}
                        for gid in remaining_counts:
                            for atlet in atlets[:]:
                                if getattr(atlet, field_name + '_id') == gid and remaining_counts[gid] > 0:
                                    result.append(atlet)
                                    atlets.remove(atlet)
                                    remaining_counts[gid] -= 1
                        return result

                    atlets_temp_main = atlets_temp_all.copy()
                    atlets_temp_pool_a = assign_atlets_to_pool(atlets_temp_main, group_counts_pool_a, group_field)
                    atlets_temp_pool_b = assign_atlets_to_pool(atlets_temp_main, group_counts_pool_b, group_field)
                    atlets_temp_pool_c = assign_atlets_to_pool(atlets_temp_main, group_counts_pool_c, group_field)
                    atlets_temp_pool_d = assign_atlets_to_pool(atlets_temp_main, group_counts_pool_d, group_field)

                    # --- Final bagan, once per category, only if it got split into pools ---
                    final_bagan = None
                    if perulangan > 1:
                        final_bagan, final_round_5 = build_full_bracket(
                            event, nomor_tanding, f'{nomor_tanding.nama_nomor_tanding} - Final', 0,
                            [], [], group_field, custom_order, atlet_assignment1
                        )

                    # --- Build each pool's bracket using the shared function ---
                    POOL_LETTERS = ['A', 'B', 'C', 'D']
                    pool_group_counts = [group_counts_pool_a, group_counts_pool_b, group_counts_pool_c, group_counts_pool_d]
                    pool_atlets = [atlets_temp_pool_a, atlets_temp_pool_b, atlets_temp_pool_c, atlets_temp_pool_d]

                    pool_round5_dbs = []

                    for i in range(1, perulangan + 1):
                        if perulangan > 1:
                            nama_bagan = f'{nomor_tanding.nama_nomor_tanding} - Pool {POOL_LETTERS[i-1]}'
                            group_counts_i = list(pool_group_counts[i-1])
                            atlets_temp_i = pool_atlets[i-1]
                        else:
                            nama_bagan = nomor_tanding.nama_nomor_tanding
                            group_counts_i = group_counts_temp
                            atlets_temp_i = atlets_temp_all

                        bagan, round_5 = build_full_bracket(
                            event, nomor_tanding, nama_bagan, perulangan,
                            group_counts_i, atlets_temp_i, group_field, custom_order, atlet_assignment1
                        )
                        pool_round5_dbs.append(round_5)

            return redirect('admin-dashboard', event_pk=event_pk)
        
        elif request.POST.get('submit_type') == 'tambah_bagan':
            nomor_tanding_pk = request.POST.get('nomor_tanding_pk')
            tipe = request.POST.get('tipe')
            if tipe == 'normal':
                return redirect('tambah-bagan', event_pk=event_pk, nomor_tanding_pk=nomor_tanding_pk)
            elif tipe == 'referchange':
                return redirect('tambah-bagan-referchange', event_pk=event_pk, nomor_tanding_pk=nomor_tanding_pk)
            elif tipe == 'round_robin':
                return redirect('tambah-bagan-round-robin', event_pk=event_pk, nomor_tanding_pk=nomor_tanding_pk)
            
    context = {
        'on': 'utama',
        'event': event,
        'admin_tatami': admin_tatami,
        'nomor_tandings': nomor_tandings,
        'bagans': bagans,
    }

    return render(request, 'admin/dashboard.html', context)

def admin_bagan_detail_round_robin(request, event_pk, bagan_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    bagan = Bagan.objects.get(pk=bagan_pk)

    all_atlets = Atlet.objects.filter(nomor_tanding=bagan.nomor_tanding).order_by('pk')
    details = DetailBagan.objects.filter(bagan=bagan)

    match_lookup = {}
    for d in details:
        pk1, pk2 = sorted([d.atlet1.pk, d.atlet2.pk])
        match_lookup[f"{pk1}-{pk2}"] = d

    table_rows = []
    for row in all_atlets:
        row_matches = []
        for col in all_atlets:
            pk1, pk2 = sorted([row.pk, col.pk])
            row_matches.append(match_lookup[f"{pk1}-{pk2}"])
        table_rows.append((row, row_matches))

    results = []
    for row_atlet, matches in table_rows:
        menang = kalah = draw = 0
        
        for match in matches:
            # Skip self matches
            if match.atlet1 == match.atlet2:
                continue
            
            if match.pemenang == '1':
                if match.atlet1 == row_atlet:
                    menang += 1
                elif match.atlet2 == row_atlet:
                    kalah += 1
            elif match.pemenang == '2':
                if match.atlet2 == row_atlet:
                    menang += 1
                elif match.atlet1 == row_atlet:
                    kalah += 1
            elif match.pemenang == '3':
                draw += 1  # Each draw counts as 1 here; you can handle 0.5 in total
            
        total = menang * 1 + draw * 0.5
        results.append({
            "atlet": row_atlet,
            "menang": menang,
            "kalah": kalah,
            "draw": draw,
            "total": total
        })

    context = {
        'on': 'utama',
        'event': event,
        'admin_tatami': admin_tatami,
        "results": results,
        'bagan': bagan,
        "all_atlets": all_atlets,
        "match_lookup": match_lookup,
        "table_rows": table_rows,
    }
    return render(request, 'admin/round-robin.html', context)

def roster_counter(request, event_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    bagans = Bagan.objects.filter(event=event).order_by('nama_bagan')

    for bagan in bagans:
        dbs = DetailBagan.objects.filter(bagan=bagan)
        bagan.count = 0
        for db in dbs:
            if db.atlet1:
                bagan.count += 1
            if db.atlet2:
                bagan.count += 1

    context = {
        'on': 'roster-counter',
        'event': event,
        'admin_tatami': admin_tatami,
        'bagans': bagans,
    }

    return render(request, 'admin/roster-counter.html', context)

def summary(request, event_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    bagans = Bagan.objects.filter(event=event).order_by('nama_bagan')

    all_utusans = Utusan.objects.filter(event=event)

    def format_rupiah(value):
        return f"{value:,}".replace(",", ".")
    
    total_counts = defaultdict(int)

    for utusan in all_utusans:
        atlet_list = utusan.atlet.all()
        utusan_totals = defaultdict(int)

        for atlet in atlet_list:
            if not atlet.nomor_tanding:
                continue 

            nama = atlet.nomor_tanding.nama_nomor_tanding.upper()

            if 'TOURNAMENT' in nama and not 'BEREGU' in nama:
                if 'PUTRA' in nama:
                    utusan_totals['tournament_putra'] += 1
                    utusan_totals['tournament_reg_putra'] += 300000
                elif 'PUTRI' in nama:
                    utusan_totals['tournament_putri'] += 1
                    utusan_totals['tournament_reg_putri'] += 300000
            elif 'TOURNAMENT' in nama and 'BEREGU' in nama and 'CS' in atlet.nama_atlet:
                utusan_totals['tournament_beregu'] += 1
                utusan_totals['tournament_reg_beregu'] += 350000

        # Assign individual totals
        utusan.total_tournament_putra = utusan_totals['tournament_putra']
        utusan.total_tournament_putri = utusan_totals['tournament_putri']
        utusan.total_tournament_beregu = utusan_totals['tournament_beregu']

        utusan.total_tournament = utusan.total_tournament_putra + utusan.total_tournament_putri + utusan.total_tournament_beregu

        utusan.total_tournament_reg_putra = utusan_totals['tournament_reg_putra']
        utusan.total_tournament_reg_putri = utusan_totals['tournament_reg_putri']
        utusan.total_tournament_reg_beregu = utusan_totals['tournament_reg_beregu']

        utusan.total_tournament_reg = utusan.total_tournament_reg_putra + utusan.total_tournament_reg_putri + utusan.total_tournament_reg_beregu

        # Update global counters
        for key, val in utusan_totals.items():
            total_counts[key] += val

        total_counts['peserta'] += utusan.total_tournament
        total_counts['reg'] += utusan.total_tournament_reg

        # Format values
        utusan.total_tournament_reg_putra = format_rupiah(utusan.total_tournament_reg_putra)
        utusan.total_tournament_reg_putri = format_rupiah(utusan.total_tournament_reg_putri)
        utusan.total_tournament_reg_beregu = format_rupiah(utusan.total_tournament_reg_beregu)
        utusan.total_tournament_reg = format_rupiah(utusan.total_tournament_reg)

    # Sort by total peserta ascending
    all_utusans = sorted(all_utusans, key=lambda u: u.total_tournament, reverse=True)

    # Format global registration counters
    formatted_totals = {
        key: format_rupiah(val)
        for key, val in total_counts.items()
        if 'reg' in key
    }

    total_biaya_kontingen = sum(
        0 for utusan in all_utusans
    )

    global_totals = {
        # Tournament
        'total_peserta_tournament_putra': total_counts['tournament_putra'],
        'total_peserta_tournament_putri': total_counts['tournament_putri'],
        'total_peserta_tournament_beregu': total_counts['tournament_beregu'],
        'total_registrasi_tournament_putra': formatted_totals.get('tournament_reg_putra', '0'),
        'total_registrasi_tournament_putri': formatted_totals.get('tournament_reg_putri', '0'),
        'total_registrasi_tournament_beregu': formatted_totals.get('tournament_reg_beregu', '0'),
        'total_peserta_tournament': total_counts['tournament_putra'] + total_counts['tournament_putri'] + total_counts['tournament_beregu'],
        'total_registrasi_tournament': format_rupiah(
            total_counts['tournament_reg_putra'] + total_counts['tournament_reg_putri'] + total_counts['tournament_reg_beregu']
        ),

        # Grand Total
        'total_peserta': total_counts['peserta'],
        'total_biaya_kontingen': format_rupiah(total_biaya_kontingen),
        'total_registrasi': format_rupiah(total_counts['reg'] + total_biaya_kontingen),
    }

    context = {
        'on': 'roster-counter',
        'event': event,
        'admin_tatami': admin_tatami,
        'bagans': bagans,
        'all_utusan': all_utusans,
        'formatted_totals': formatted_totals,
        'global_totals': global_totals,
    }

    return render(request, 'admin/summary.html', context)

def admin_bagan_detail(request, event_pk, bagan_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    tatami = admin_tatami.tatami
    bagan = Bagan.objects.get(pk=bagan_pk)
    if bagan.round_robin:
        return redirect('admin-bagan-detail-round-robin', event_pk=event_pk, bagan_pk=bagan_pk)
    all_atlets = Atlet.objects.filter(nomor_tanding=bagan.nomor_tanding)
    detail_bagans_round_1 = DetailBagan.objects.filter(bagan=bagan, round=1).order_by('urutan')
    detail_bagans_round_2 = DetailBagan.objects.filter(bagan=bagan, round=2).order_by('urutan')
    detail_bagans_round_3 = DetailBagan.objects.filter(bagan=bagan, round=3).order_by('urutan')
    detail_bagans_round_4 = DetailBagan.objects.filter(bagan=bagan, round=4).order_by('urutan')
    detail_bagan_round_5 = DetailBagan.objects.filter(bagan=bagan, round=5).first()


    if 'REFERCHANGE' in bagan.nama_bagan:
        referchange = True
    else:
        referchange = False

    if request.method == 'POST':
        if request.POST.get('submit_type') == 'simpan_juara':
            juara_1_pk = request.POST.get('juara_1_pk')
            juara_2_pk = request.POST.get('juara_2_pk')
            juara_3a_pk = request.POST.get('juara_3a_pk')
            juara_3b_pk = request.POST.get('juara_3b_pk')
            if juara_1_pk == '-':
                bagan.juara_1 = None
            else:
                bagan.juara_1 = Atlet.objects.filter(pk=juara_1_pk).first()
            if juara_2_pk == '-':
                bagan.juara_2 = None
            else:
                bagan.juara_2 = Atlet.objects.filter(pk=juara_2_pk).first()
            if juara_3a_pk == '-':
                bagan.juara_3a = None
            else:
                bagan.juara_3a = Atlet.objects.filter(pk=juara_3a_pk).first()
            if juara_3b_pk == '-':
                bagan.juara_3b = None
            else:
                bagan.juara_3b = Atlet.objects.filter(pk=juara_3b_pk).first()
        elif request.POST.get('submit_type') == 'generate_juara':
            if detail_bagan_round_5.atlet1:
                bagan.juara_1 = detail_bagan_round_5.atlet1
            else:
                bagan.juara_1 = None
            for db in detail_bagans_round_4:
                if db.pemenang == '1':
                    bagan.juara_2 = db.atlet2
                elif db.pemenang == '2':
                    bagan.juara_2 = db.atlet1
                else:
                    bagan.juara_2 = None
            for i, db in enumerate(detail_bagans_round_3):
                if db.pemenang == '1':
                    pemenang = db.atlet2
                elif db.pemenang == '2':
                    pemenang = db.atlet1
                else:
                    pemenang = None
                if i == 0:
                    bagan.juara_3a = pemenang
                else:
                    bagan.juara_3b = pemenang
        bagan.save()

        payload = {
            'status': 'finished',
            'kode_realtime': f'{detail_bagan_round_5.bagan.pk}-{detail_bagan_round_5.pk}',
            'juara_1': bagan.juara_1.nama_atlet if bagan.juara_1 else None,
            'juara_2': bagan.juara_2.nama_atlet if bagan.juara_2 else None,
            'juara_3a': bagan.juara_3a.nama_atlet if bagan.juara_3a else None,
            'juara_3b': bagan.juara_3b.nama_atlet if bagan.juara_3b else None,
        }
        success, result = send_to_hosted(payload, endpoint='api/final-result/')

        if not success:
            messages.warning(
                request,
                f'Hasil berhasil disimpan secara lokal, tapi gagal mengirim ke server: {result}'
            )

        return redirect('admin-bagan-detail', event_pk=event_pk, bagan_pk=bagan_pk)

    context = {
        'on': 'utama',
        'event': event,
        'admin_tatami': admin_tatami,
        'tatami': tatami,
        'bagan': bagan,
        'detail_bagans_round_1': detail_bagans_round_1,
        'detail_bagans_round_2': detail_bagans_round_2,
        'detail_bagans_round_3': detail_bagans_round_3,
        'detail_bagans_round_4': detail_bagans_round_4,
        'detail_bagan_round_5': detail_bagan_round_5,
        'all_atlets': all_atlets,
        'referchange': referchange,
    }

    return render(request, 'admin/bagan-detail.html', context)

def tambah_bagan(request, event_pk, nomor_tanding_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    nomor_tanding = NomorTanding.objects.filter(pk=nomor_tanding_pk).first()
    all_atlets = Atlet.objects.filter(nomor_tanding=nomor_tanding)

    round_1 = [1, 2, 3, 4, 5, 6, 7, 8]
    round_2 = [1, 2, 3, 4]
    round_3 = [1, 2]
    round_4 = [1]

    if request.method == 'POST':
        if request.POST.get('submit_type') == 'simpan-bagan':
            rounds_data = [
                (1, request.POST.getlist('atlet_round_1_aka_pk'), request.POST.getlist('atlet_round_1_ao_pk')),
                (2, request.POST.getlist('atlet_round_2_aka_pk'), request.POST.getlist('atlet_round_2_ao_pk')),
                (3, request.POST.getlist('atlet_round_3_aka_pk'), request.POST.getlist('atlet_round_3_ao_pk')),
                (4, request.POST.getlist('atlet_round_4_aka_pk'), request.POST.getlist('atlet_round_4_ao_pk')),
            ]
            nama_bagan = request.POST.get('nama_bagan').strip().upper()

            if 'KATA' in nomor_tanding.nama_nomor_tanding:
                tipe_tanding = '1'
            else:
                tipe_tanding = '2'

            new_bagan = Bagan.objects.create(event=event, nomor_tanding=nomor_tanding, tipe_tanding=tipe_tanding, nama_bagan=nama_bagan)

            for round_number, aka_list, ao_list in rounds_data:
                for index, (aka_pk, ao_pk) in enumerate(zip(aka_list, ao_list), start=1):
                    atlet1 = Atlet.objects.filter(pk=aka_pk).first() if aka_pk != '-' else None
                    atlet2 = Atlet.objects.filter(pk=ao_pk).first() if ao_pk != '-' else None

                    DetailBagan.objects.create(
                        bagan=new_bagan,
                        round=round_number,
                        urutan=index,
                        atlet1=atlet1,
                        atlet2=atlet2
                    )

            round_5 = DetailBagan.objects.create(bagan=new_bagan, round=5, urutan=1)
                

            messages.success(request, f'Berhasil membuat bagan: {nama_bagan}')
            return redirect('admin-dashboard', event_pk=event_pk)

    context = {
        'on': 'utama',
        'event': event,
        'admin_tatami': admin_tatami,
        'nomor_tanding': nomor_tanding,
        'round_1': round_1,
        'round_2': round_2,
        'round_3': round_3,
        'round_4': round_4,
        'all_atlets': all_atlets,
    }

    return render(request, 'admin/tambah-bagan.html', context)

def tambah_bagan_referchange(request, event_pk, nomor_tanding_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    nomor_tanding = NomorTanding.objects.filter(pk=nomor_tanding_pk).first()
    all_atlets = Atlet.objects.filter(nomor_tanding=nomor_tanding)

    round_1 = [1]
    round_2 = [1]
    round_3 = [1]

    if request.method == 'POST':
        if request.POST.get('submit_type') == 'simpan-bagan':
            rounds_data = [
                (1, request.POST.getlist('atlet_round_1_aka_pk'), request.POST.getlist('atlet_round_1_ao_pk')),
                (2, request.POST.getlist('atlet_round_2_aka_pk'), request.POST.getlist('atlet_round_2_ao_pk')),
                (3, request.POST.getlist('atlet_round_3_aka_pk'), request.POST.getlist('atlet_round_3_ao_pk')),
            ]
            nama_bagan = request.POST.get('nama_bagan').strip().upper()

            if 'KATA' in nomor_tanding.nama_nomor_tanding:
                tipe_tanding = '1'
            else:
                tipe_tanding = '2'

            new_bagan = Bagan.objects.create(event=event, nomor_tanding=nomor_tanding, tipe_tanding=tipe_tanding, nama_bagan=nama_bagan)

            for round_number, aka_list, ao_list in rounds_data:
                for index, (aka_pk, ao_pk) in enumerate(zip(aka_list, ao_list), start=1):
                    atlet1 = Atlet.objects.filter(pk=aka_pk).first() if aka_pk != '-' else None
                    atlet2 = Atlet.objects.filter(pk=ao_pk).first() if ao_pk != '-' else None

                    DetailBagan.objects.create(
                        bagan=new_bagan,
                        round=round_number,
                        urutan=index,
                        atlet1=atlet1,
                        atlet2=atlet2
                    )

            round_4 = DetailBagan.objects.create(bagan=new_bagan, round=4, urutan=1)

            return redirect('admin-dashboard', event_pk=event_pk)

    context = {
        'on': 'utama',
        'event': event,
        'admin_tatami': admin_tatami,
        'nomor_tanding': nomor_tanding,
        'round_1': round_1,
        'round_2': round_2,
        'round_3': round_3,
        'all_atlets': all_atlets,
    }

    return render(request, 'admin/tambah-bagan-referchange.html', context)

def tambah_bagan_round_robin(request, event_pk, nomor_tanding_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    nomor_tanding = NomorTanding.objects.filter(pk=nomor_tanding_pk).first()
    all_atlets = list(
        Atlet.objects.filter(nomor_tanding=nomor_tanding).order_by('pk')
    )

    if 'KATA' in nomor_tanding.nama_nomor_tanding:
        tipe_tanding = '1'
    else:
        tipe_tanding = '2'

    new_bagan = Bagan.objects.create(event=event, nama_bagan=f'ROUND ROBIN {nomor_tanding.nama_nomor_tanding}', nomor_tanding=nomor_tanding, tipe_tanding=tipe_tanding, round_robin=True)
    
    match_lookup = {}
    for atlet_1 in all_atlets:
        for atlet_2 in all_atlets:
            key = tuple(sorted([atlet_1.pk, atlet_2.pk]))
            if key not in match_lookup:
                match = DetailBagan.objects.create(
                    bagan=new_bagan,
                    atlet1=atlet_1,
                    atlet2=atlet_2
                )
                match_lookup[key] = match

    return redirect('admin-dashboard', event_pk=event_pk)

def edit_admin_bagan_detail(request, event_pk, bagan_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    bagan = Bagan.objects.get(pk=bagan_pk)
    all_atlets = Atlet.objects.filter(nomor_tanding=bagan.nomor_tanding)
    detail_bagans_round_1 = DetailBagan.objects.filter(bagan=bagan, round=1).order_by('urutan')
    detail_bagans_round_2 = DetailBagan.objects.filter(bagan=bagan, round=2).order_by('urutan')
    detail_bagans_round_3 = DetailBagan.objects.filter(bagan=bagan, round=3).order_by('urutan')
    detail_bagans_round_4 = DetailBagan.objects.filter(bagan=bagan, round=4).order_by('urutan')
    detail_bagan_round_5 = DetailBagan.objects.filter(bagan=bagan, round=5).first()

    if request.method == 'POST':
        if request.POST.get('submit_type') == 'simpan_nama':
            nama_bagan = request.POST.get('nama_bagan')
            if nama_bagan:
                bagan.nama_bagan = nama_bagan
                bagan.save()

        return redirect('edit-admin-bagan-detail', event_pk=event_pk, bagan_pk=bagan_pk)

    context = {
        'on': 'utama',
        'event': event,
        'admin_tatami': admin_tatami,
        'bagan': bagan,
        'detail_bagans_round_1': detail_bagans_round_1,
        'detail_bagans_round_2': detail_bagans_round_2,
        'detail_bagans_round_3': detail_bagans_round_3,
        'detail_bagans_round_4': detail_bagans_round_4,
        'detail_bagan_round_5': detail_bagan_round_5,
        'all_atlets': all_atlets,
    }

    return render(request, 'admin/edit-bagan-detail.html', context)

def hapus_admin_bagan_detail(request, event_pk, bagan_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    bagan = Bagan.objects.get(pk=bagan_pk)
    bagan.delete()

    messages.success(request, f'Berhasil menghapus bagan: {bagan.nama_bagan}')
    return redirect('admin-dashboard', event_pk=event_pk)

def admin_edit_detail_bagan(request, event_pk, bagan_pk, detailbagan_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    bagan = Bagan.objects.get(pk=bagan_pk)
    detail_bagan = DetailBagan.objects.get(pk=detailbagan_pk)
    atlets = Atlet.objects.filter(nomor_tanding=bagan.nomor_tanding)

    if request.method == 'POST':
        if request.POST.get('submit_type') == 'atlet-simpan':
            atlet_1_pk = request.POST.get('atlet-aka')
            if atlet_1_pk != '-':
                detail_bagan.atlet1 = Atlet.objects.filter(pk=atlet_1_pk).first()
            else:
                detail_bagan.atlet1 = None
            atlet_2_pk = request.POST.get('atlet-ao')
            if atlet_2_pk != '-':
                detail_bagan.atlet2 = Atlet.objects.filter(pk=atlet_2_pk).first()
            else:
                detail_bagan.atlet2 = None
            detail_bagan.save()

            payload = {
                'status': 'edit',
                'kode_realtime': f'{detail_bagan.bagan.pk}-{detail_bagan.pk}',
                'atlet_aka': detail_bagan.atlet1.nama_atlet if detail_bagan.atlet1 else None,
                'atlet_ao': detail_bagan.atlet2.nama_atlet if detail_bagan.atlet2 else None,
                'utusan_aka': detail_bagan.atlet1.utusan.nama_utusan if detail_bagan.atlet1 else None,
                'utusan_ao': detail_bagan.atlet2.utusan.nama_utusan if detail_bagan.atlet2 else None,
            }
            success, result = send_to_hosted(payload, endpoint='api/edit-bagan/')
    
            if not success:
                messages.warning(
                    request,
                    f'Hasil berhasil disimpan secara lokal, tapi gagal mengirim ke server: {result}'
                )
        
        return redirect('edit-detail-bagan', event_pk=event_pk, bagan_pk=bagan_pk, detailbagan_pk=detailbagan_pk)

    context = {
        'on': 'utama',
        'event': event,
        'admin_tatami': admin_tatami,
        'bagan': bagan,
        'detail_bagan': detail_bagan,
        'atlets': atlets,
    }

    return render(request, 'admin/edit-detail-bagan.html', context)

def control_panel(request, event_pk, bagan_pk, detailbagan_pk, tatami_pk):
    event = Event.objects.get(pk=event_pk)
    tatami = Tatami.objects.get(pk=tatami_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    bagan = Bagan.objects.get(pk=bagan_pk)
    detail_bagan = DetailBagan.objects.get(pk=detailbagan_pk)
    aka_score_obj = Score.objects.filter(detail_bagan=detail_bagan, atlet=0).first()
    ao_score_obj = Score.objects.filter(detail_bagan=detail_bagan, atlet=1).first()

    detail_bagan.save()

    if not aka_score_obj:
        aka_score_obj = Score.objects.create(detail_bagan=detail_bagan, atlet=0)
    if not ao_score_obj:
        ao_score_obj = Score.objects.create(detail_bagan=detail_bagan, atlet=1)

    tatami = admin_tatami.tatami
    tatami.detail_bagan = detail_bagan
    tatami.save()

    total_aka_score = 0
    total_ao_score = 0

    mu = Matchup.objects.filter(db=detail_bagan).first()

    if 'KUMITE BEREGU' in bagan.nama_bagan and mu:
        for matchup in Matchup.objects.filter(bagan=bagan, detail_bagan=mu.detail_bagan):
            if matchup.db.pemenang == '1':
                total_aka_score += 1
            elif matchup.db.pemenang == '2':
                total_ao_score += 1

    if request.method == 'POST':
        pemenang = request.POST.get('pemenang')
        if request.POST.get('submit_type') == 'kata-simpan':
            aka_scores = request.POST.getlist('akaScores')
            ao_scores = request.POST.getlist('aoScores')
            total_aka = request.POST.get('totalAka')
            total_ao = request.POST.get('totalAo')
            kata_aka = request.POST.get('kata-aka')
            kata_ao = request.POST.get('kata-ao')

            score_fields = ['score1', 'score2', 'score3', 'score4', 'score5']
        
            for i, field in enumerate(score_fields):
                if i < len(aka_scores):
                    setattr(aka_score_obj, field, aka_scores[i])
            
            for i, field in enumerate(score_fields):
                if i < len(ao_scores):
                    setattr(ao_score_obj, field, ao_scores[i])
            
            aka_score_obj.save()
            ao_score_obj.save()

            detail_bagan.score1 = total_aka
            detail_bagan.score2 = total_ao
            detail_bagan.kata1 = kata_aka
            detail_bagan.kata2 = kata_ao
        
        elif request.POST.get('submit_type') == 'kumite-simpan':
            aka_score = request.POST.get('akaScore')
            ao_score = request.POST.get('aoScore')
            aka_vr = bool(request.POST.get('aka-vr'))
            ao_vr = bool(request.POST.get('ao-vr'))

            detail_bagan.score1 = aka_score
            detail_bagan.score2 = ao_score

            detail_bagan.vr1 = aka_vr
            detail_bagan.vr2 = ao_vr
        
        if pemenang == 'aka':
            detail_bagan.pemenang = '1'
        elif pemenang == 'ao':
            detail_bagan.pemenang = '2'
        else:
            detail_bagan.pemenang = '3'

        detail_bagan.selesai = True
        detail_bagan.save()

        winner_atlet = None
        
        if not bagan.round_robin or not detail_bagan.team:
            next_round_number = detail_bagan.round + 1
            next_round_urutan = (detail_bagan.urutan + 1) // 2
            detailbagan_next_round = DetailBagan.objects.filter(bagan=bagan, round=next_round_number, urutan=next_round_urutan).first()

            if detailbagan_next_round:
                if pemenang == 'aka':
                    winner_atlet = detail_bagan.atlet1
                    detail_bagan.pemenang = '1'
                elif pemenang == 'ao':
                    winner_atlet = detail_bagan.atlet2
                    detail_bagan.pemenang = '2'
                else:
                    winner_atlet = None
                    detail_bagan.pemenang = '3'

                if winner_atlet:
                    if detail_bagan.urutan % 2 == 1:
                        detailbagan_next_round.atlet1 = winner_atlet
                        if detail_bagan.vr1 and pemenang == 'aka':
                            detailbagan_next_round.vr1 = True
                        elif detail_bagan.vr2 and pemenang == 'ao':
                            detailbagan_next_round.vr1 = True
                    else:
                        detailbagan_next_round.atlet2 = winner_atlet
                        if detail_bagan.vr1 and pemenang == 'aka':
                            detailbagan_next_round.vr2 = True
                        elif detail_bagan.vr2 and pemenang == 'ao':
                            detailbagan_next_round.vr2 = True

                detail_bagan.save()
                detailbagan_next_round.save()

        if detail_bagan.round != 10:
            payload = {
                'status': 'finished',
                'round': detail_bagan.round,
                'urutan': detail_bagan.urutan,
                'kode_realtime': f'{detail_bagan.bagan.pk}-{detail_bagan.pk}',
                'pemenang': pemenang,
                'score_aka': detail_bagan.score1,
                'score_ao': detail_bagan.score2,
                'vr1': detail_bagan.vr1,
                'vr2': detail_bagan.vr2,
                'next_vr1': detailbagan_next_round.vr1,
                'next_vr2': detailbagan_next_round.vr2,
                'winner_atlet': winner_atlet.nama_atlet if winner_atlet else None,
                'next_kode_realtime': f'{detailbagan_next_round.bagan.pk}-{detailbagan_next_round.pk}',
                'ring_number': Tatami.objects.filter(detail_bagan=detail_bagan).first().tatami_number,
            }
            success, result = send_to_hosted(payload, endpoint='api/result/')

            if not success:
                messages.warning(
                    request,
                    f'Hasil berhasil disimpan secara lokal, tapi gagal mengirim ke server: {result}'
                )

        return redirect('admin-bagan-detail', event_pk=event_pk, bagan_pk=bagan_pk)

    detail_data = {
        "atlet_red": detail_bagan.atlet1.nama_atlet if detail_bagan.atlet1 else None,
        "atlet_red_perguruan": detail_bagan.atlet1.perguruan.nama_perguruan if detail_bagan.atlet1 else None,
        "atlet_red_utusan": detail_bagan.atlet1.utusan.nama_utusan if detail_bagan.atlet1 else None,
        "atlet_red_kata": detail_bagan.kata1 if detail_bagan.kata1 else None,
        "atlet_red_vr": detail_bagan.vr1 if detail_bagan.vr1 else None,
        "atlet_blue": detail_bagan.atlet2.nama_atlet if detail_bagan.atlet2 else None,
        "atlet_blue_perguruan": detail_bagan.atlet2.perguruan.nama_perguruan if detail_bagan.atlet2 else None,
        "atlet_blue_utusan": detail_bagan.atlet2.utusan.nama_utusan if detail_bagan.atlet2 else None,
        "atlet_blue_kata": detail_bagan.kata2 if detail_bagan.kata2 else None,
        "atlet_blue_vr": detail_bagan.vr2 if detail_bagan.vr2 else None,
        "tipe_tanding": bagan.tipe_tanding,
        "team": True if 'KUMITE BEREGU' in bagan.nama_bagan else None,
        "total_aka_score": total_aka_score,
        "total_ao_score": total_ao_score,
        "nomor_tanding": bagan.nomor_tanding.nama_nomor_tanding,
    } 

    group_name = f"scoring_{admin_tatami.tatami.pk}"
    channel_layer = get_channel_layer()

    async_to_sync(channel_layer.group_send)(
        group_name,
        {
            "type": "broadcast_command",
            "message": "get_atlet",
            "details": detail_data,
        }
    )

    group_name = f"juryroom_{admin_tatami.tatami.pk}"
    channel_layer = get_channel_layer()

    async_to_sync(channel_layer.group_send)(
        group_name,
        {
            "type": "broadcast_command",
            "message": "get_atlet",
            "details": detail_data,
        }
    )

    group_name = f"coachroom_{admin_tatami.tatami.pk}"
    channel_layer = get_channel_layer()

    async_to_sync(channel_layer.group_send)(
        group_name,
        {
            "type": "broadcast_command",
            "message": "get_atlet",
            "details": [detail_bagan.vr1, detail_bagan.vr2],
        }
    )

    context = {
        'on': 'utama',
        'event': event,
        'admin_tatami': admin_tatami,
        'bagan': bagan,
        'detail_bagan': detail_bagan,
        'aka_score': aka_score_obj,
        'ao_score': ao_score_obj,
        'tatami': tatami,
    }

    return render(request, 'admin/control-panel.html', context)

def control_panel_fest(request, event_pk, tatami_pk):
    event = Event.objects.get(pk=event_pk)
    tatami = Tatami.objects.get(pk=tatami_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()

    total_aka_score = 0
    total_ao_score = 0

    detail_data = {
        "atlet_red": "Aka",
        "atlet_red_perguruan": "-",
        "atlet_red_utusan": "-",
        "atlet_red_kata": "-",
        "atlet_red_vr": None,
        "atlet_blue": "Ao",
        "atlet_blue_perguruan": "-",
        "atlet_blue_utusan": "-",
        "atlet_blue_kata": "-",
        "atlet_blue_vr": None,
        "tipe_tanding": "2",
        "team": None,
        "total_aka_score": total_aka_score,
        "total_ao_score": total_ao_score,
        "nomor_tanding": "Festival",
    } 

    group_name = f"scoring_{admin_tatami.tatami.pk}"
    channel_layer = get_channel_layer()

    async_to_sync(channel_layer.group_send)(
        group_name,
        {
            "type": "broadcast_command",
            "message": "get_atlet",
            "details": detail_data,
        }
    )

    context = {
        'on': 'fest',
        'event': event,
        'admin_tatami': admin_tatami,
        'tatami': tatami,
    }

    return render(request, 'admin/control-panel-fest.html', context)

def control_panel_team(request, event_pk, bagan_pk, detailbagan_pk, tatami_pk):
    event = Event.objects.get(pk=event_pk)
    tatami = Tatami.objects.get(pk=tatami_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    bagan = Bagan.objects.get(pk=bagan_pk)
    detail_bagan = DetailBagan.objects.get(pk=detailbagan_pk)

    tatami = admin_tatami.tatami
    tatami.detail_bagan = detail_bagan
    tatami.save()

    team_aka = Atlet.objects.filter(utusan=detail_bagan.atlet1.utusan, nomor_tanding=detail_bagan.atlet1.nomor_tanding).exclude(nama_atlet__icontains='team')
    team_ao = Atlet.objects.filter(utusan=detail_bagan.atlet2.utusan, nomor_tanding=detail_bagan.atlet2.nomor_tanding).exclude(nama_atlet__icontains='team')

    matchups = Matchup.objects.filter(bagan=bagan, detail_bagan=detail_bagan).order_by('round')
    team_aka_score = 0
    team_ao_score = 0 
    team_aka_lil_score = 0
    team_ao_lil_score = 0

    for matchup in matchups:
        team_aka_lil_score += int(matchup.db.score1) if matchup.db.score1 else 0
        team_ao_lil_score += int(matchup.db.score2) if matchup.db.score2 else 0
        if matchup.db.pemenang == '1':
            team_aka_score += 1
        elif matchup.db.pemenang == '2':
            team_ao_score += 1 
    
    if request.method == 'POST':
        if request.POST.get('submit_type') == 'save':
            i = 1
            while True:
                aka = request.POST.get(f'aka_{i}')
                ao = request.POST.get(f'ao_{i}')
                if aka is None:
                    break
                if aka == '-' or ao == '-':
                    i += 1
                    continue
                aka_atlet = Atlet.objects.get(pk=aka)
                ao_atlet = Atlet.objects.get(pk=ao)
                already_exists = Matchup.objects.filter(
                    bagan=bagan,
                    detail_bagan=detail_bagan,
                    round=i
                ).exists()

                if not already_exists:
                    new_detail_bagan = DetailBagan.objects.create(
                        bagan=bagan,
                        round=10,
                        urutan=1,
                        atlet1=aka_atlet,
                        atlet2=ao_atlet,
                        score1=0,
                        score2=0,
                        vr1=True,
                        vr2=True,
                        team=True,
                    )
                    Matchup.objects.create(
                        bagan=bagan,
                        detail_bagan=detail_bagan,
                        db=new_detail_bagan,
                        round=i,
                    )

                i += 1
        elif request.POST.get('submit_type') == 'delete':
            matchup_pk = request.POST.get('matchup')
            matchup = Matchup.objects.get(pk=matchup_pk)
            matchup.db.delete()
            matchup.delete()
        
        elif request.POST.get('submit_type') == 'simpan':
            if team_aka_score > team_ao_score:
                detail_bagan.pemenang = '1'
            elif team_ao_score > team_aka_score:
                detail_bagan.pemenang = '2'
            elif team_aka_lil_score > team_ao_lil_score:
                detail_bagan.pemenang = '1'
            elif team_ao_lil_score > team_aka_lil_score:
                detail_bagan.pemenang = '2'
            else:
                detail_bagan.pemenang = '3'


            detail_bagan.score1 = team_aka_score
            detail_bagan.score2 = team_ao_score
            detail_bagan.scorekecil1 = team_aka_lil_score
            detail_bagan.scorekecil2 = team_ao_lil_score
            detail_bagan.selesai = True
            detail_bagan.save()
            
            next_round_number = detail_bagan.round + 1
            next_round_urutan = (detail_bagan.urutan + 1) // 2
            detailbagan_next_round = DetailBagan.objects.filter(bagan=bagan, round=next_round_number, urutan=next_round_urutan).first()

            if detailbagan_next_round:
                if team_aka_score > team_ao_score:
                    winner_atlet = detail_bagan.atlet1
                elif team_ao_score > team_aka_score:
                    winner_atlet = detail_bagan.atlet2
                elif team_aka_lil_score > team_ao_lil_score:
                    winner_atlet = detail_bagan.atlet1
                elif team_ao_lil_score > team_aka_lil_score:
                    winner_atlet = detail_bagan.atlet2
                else:
                    winner_atlet = None
                    detail_bagan.pemenang = '3'

                if winner_atlet:
                    if detail_bagan.urutan % 2 == 1:
                        detailbagan_next_round.atlet1 = winner_atlet
                        if detail_bagan.vr1 and detail_bagan.pemenang == '1':
                            detailbagan_next_round.vr1 = True
                        elif detail_bagan.vr2 and detail_bagan.pemenang == '2':
                            detailbagan_next_round.vr1 = True
                    else:
                        detailbagan_next_round.atlet2 = winner_atlet
                        if detail_bagan.vr1 and detail_bagan.pemenang == '1':
                            detailbagan_next_round.vr2 = True
                        elif detail_bagan.vr2 and detail_bagan.pemenang == '2':
                            detailbagan_next_round.vr2 = True

                detail_bagan.save()
                detailbagan_next_round.save()

                if winner_atlet == detail_bagan.atlet1:
                    pemenang = 'aka'
                elif winner_atlet == detail_bagan.atlet2:
                    pemenang = 'ao'
                else:
                    pemenang = '3'

                payload = {
                    'status': 'finished',
                    'pemenang': pemenang,
                    'round': detail_bagan.round,
                    'urutan': detail_bagan.urutan,
                    'kode_realtime': f'{detail_bagan.bagan.pk}-{detail_bagan.pk}',
                    'score_aka': detail_bagan.score1,
                    'score_ao': detail_bagan.score2,
                    'lil_score_aka': detail_bagan.scorekecil1,
                    'lil_score_ao': detail_bagan.scorekecil2,
                    'vr1': detail_bagan.vr1,
                    'vr2': detail_bagan.vr2,
                    'next_vr1': detailbagan_next_round.vr1,
                    'next_vr2': detailbagan_next_round.vr2,
                    'winner_atlet': winner_atlet.nama_atlet if winner_atlet else None,
                    'next_kode_realtime': f'{detailbagan_next_round.bagan.pk}-{detailbagan_next_round.pk}',
                    'ring_number': Tatami.objects.filter(detail_bagan=detail_bagan).first().tatami_number,
                }
                success, result = send_to_hosted(payload, endpoint='api/result/')

                if not success:
                    messages.warning(
                        request,
                        f'Hasil berhasil disimpan secara lokal, tapi gagal mengirim ke server: {result}'
                    )

            return redirect('admin-bagan-detail', event_pk=event_pk, bagan_pk=bagan_pk)
            
        return redirect("control-panel-team", event_pk=event_pk, bagan_pk=bagan_pk, detailbagan_pk=detailbagan_pk, tatami_pk=tatami_pk)

    context = {
        'on': 'utama',
        'event': event,
        'admin_tatami': admin_tatami,
        'bagan': bagan,
        'detail_bagan': detail_bagan,
        'tatami': tatami,
        'team_aka': team_aka,
        'team_ao': team_ao,
        'matchups': matchups,
        'team_aka_score': team_aka_score,
        'team_ao_score': team_ao_score,
        'team_aka_lil_score': team_aka_lil_score,
        'team_ao_lil_score': team_ao_lil_score,
    }

    return render(request, 'admin/control-panel-team.html', context)

@csrf_exempt
def message_retriever(request, tatami_pk):
    if request.method == 'POST':
        action = request.POST.get('action')
        details = request.POST.get('details')
        tatami = Tatami.objects.get(pk=tatami_pk)

        group_name = f"scoring_{tatami.pk}"
        channel_layer = get_channel_layer()

        async_to_sync(channel_layer.group_send)(
            group_name,
            {
                "type": "broadcast_command",
                "message": action,
                "details": details,
            }
        )

        return JsonResponse({'status': 'ok'})
    return JsonResponse({'error': 'Invalid method'}, status=405)

def admin_atlet(request, event_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    perguruans = Perguruan.objects.filter(event=event)
    utusans = Utusan.objects.filter(event=event)
    nomor_tandings = NomorTanding.objects.filter(event=event)
    atlets = Atlet.objects.filter(event=event).order_by('-pk')

    if request.method == 'POST':
        if request.POST.get('submit_type') == 'import_atlet':
            excel_file = request.FILES.get('excel_atlet')

            if not excel_file:
                messages.error(request, "Silakan pilih file Excel.")
                return redirect('admin-atlet', event_pk=event_pk)
            
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    nama, perguruan, utusan, nomor_tanding = row

                    perguruan_name = perguruan.strip().upper()
                    perguruan_obj, _ = Perguruan.objects.get_or_create(event=event, nama_perguruan=perguruan_name)

                    utusan_name = utusan.strip().upper()
                    utusan_obj, _ = Utusan.objects.get_or_create(event=event, nama_utusan=utusan_name)

                    nomor_tanding_name = nomor_tanding.strip().upper()
                    nomor_tanding_obj, _ = NomorTanding.objects.get_or_create(event=event, nama_nomor_tanding=nomor_tanding_name)

                    Atlet.objects.create(
                        event=event,
                        nama_atlet=nama,
                        perguruan=perguruan_obj,
                        utusan=utusan_obj,
                        nomor_tanding=nomor_tanding_obj
                    )

                messages.success(request, "Data atlet berhasil diimport.")
            except Exception as e:
                messages.error(request, f"Gagal mengimport file: {str(e)}")
        
        elif request.POST.get('submit_type') == 'tambah_atlet':
            nik = request.POST.get('nik', '').strip()
            Atlet.objects.create(
                event=event,
                nama_atlet=request.POST.get('nama_atlet', '').strip().upper(),
                nik=nik,
                perguruan_id=request.POST.get('perguruan') or None,
                utusan_id=request.POST.get('utusan') or None,
                nomor_tanding_id=request.POST.get('nomor_tanding') or None,
            )

            messages.success(request, "Berhasil menambahkan atlet.")
            
        return redirect('admin-atlet', event_pk=event_pk)

    context = {
        'on': 'atlet',
        'event': event,
        'admin_tatami': admin_tatami,
        'atlets': atlets,
        'perguruans': perguruans,
        'utusans': utusans,
        'nomor_tandings': nomor_tandings,
    }

    return render(request, 'admin/atlet.html', context)

@require_POST
def edit_atlet_ajax(request):
    atlet_id = request.POST.get('atlet_id')
    atlet = Atlet.objects.filter(pk=atlet_id).first()
    if not atlet:
        return JsonResponse({'success': False, 'message': 'Atlet tidak ditemukan.'}, status=404)

    nama_atlet = request.POST.get('nama_atlet', '').strip().upper()
    if not nama_atlet:
        return JsonResponse({'success': False, 'message': 'Nama atlet wajib diisi.'}, status=400)

    atlet.nama_atlet = nama_atlet
    atlet.nik = request.POST.get('nik', '').strip()
    atlet.perguruan_id = request.POST.get('perguruan') or None
    atlet.utusan_id = request.POST.get('utusan') or None
    atlet.nomor_tanding_id = request.POST.get('nomor_tanding') or None
    atlet.save()

    return JsonResponse({
        'success': True,
        'atlet_id': atlet.pk,
        'nama_atlet': atlet.nama_atlet,
        'nik': atlet.nik,
        'perguruan_id': atlet.perguruan_id or '',
        'perguruan_nama': atlet.perguruan.nama_perguruan if atlet.perguruan else '',
        'utusan_id': atlet.utusan_id or '',
        'utusan_nama': atlet.utusan.nama_utusan if atlet.utusan else '',
        'nomor_tanding_id': atlet.nomor_tanding_id or '',
        'nomor_tanding_nama': atlet.nomor_tanding.nama_nomor_tanding if atlet.nomor_tanding else '',
    })

def get_atlet_nik(request, atlet_pk):
    atlet = get_object_or_404(Atlet, pk=atlet_pk)
    return JsonResponse({'nik': atlet.nik})

def admin_nomor_tanding(request, event_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    nomor_tandings = NomorTanding.objects.filter(event=event).order_by('-pk')

    if request.method == 'POST':
        if request.POST.get('submit_type') == 'tambah_nomor_tanding':
            nama_nomor_tanding = request.POST.get('nomor_tanding').strip().upper()
            new_nomor_tanding = NomorTanding.objects.create(event=event, nama_nomor_tanding=nama_nomor_tanding)
        elif request.POST.get('submit_type') == 'hapus':
            nomor_tanding_pk = request.POST.get('nomor_tanding_pk')
            NomorTanding.objects.get(pk=nomor_tanding_pk).delete()

        return redirect('admin-nomor-tanding', event_pk=event_pk)
    
    context = {
        'on': 'nomor-tanding',
        'event': event,
        'admin_tatami': admin_tatami,
        'nomor_tandings': nomor_tandings,
    }
    return render(request, 'admin/nomor-tanding.html', context)

def admin_utusan(request, event_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    utusans = Utusan.objects.filter(event=event)
    
    utusan_medals = defaultdict(lambda: {"gold": 0, "silver": 0, "bronze": 0})
    utusan_winners = []

    bagans = Bagan.objects.filter(event=event)

    for bagan in bagans:
        if bagan.juara_1 and bagan.juara_1.utusan:
            utusan_medals[bagan.juara_1.utusan.pk]["gold"] += 1
            utusan_winners.append(({"pk": bagan.juara_1.utusan.pk, "nama_atlet": bagan.juara_1.nama_atlet, "perguruan": bagan.juara_1.perguruan.nama_perguruan, "juara": "1", "nama_bagan": bagan.nama_bagan}))
        if bagan.juara_2 and bagan.juara_2.utusan:
            utusan_medals[bagan.juara_2.utusan.pk]["silver"] += 1
            utusan_winners.append(({"pk": bagan.juara_2.utusan.pk, "nama_atlet": bagan.juara_2.nama_atlet, "perguruan": bagan.juara_2.perguruan.nama_perguruan, "juara": "2", "nama_bagan": bagan.nama_bagan}))
        if bagan.juara_3a and bagan.juara_3a.utusan:
            utusan_medals[bagan.juara_3a.utusan.pk]["bronze"] += 1
            utusan_winners.append(({"pk": bagan.juara_3a.utusan.pk, "nama_atlet": bagan.juara_3a.nama_atlet, "perguruan": bagan.juara_3a.perguruan.nama_perguruan, "juara": "3a", "nama_bagan": bagan.nama_bagan}))
        if bagan.juara_3b and bagan.juara_3b.utusan:
            utusan_medals[bagan.juara_3b.utusan.pk]["bronze"] += 1
            utusan_winners.append(({"pk": bagan.juara_3b.utusan.pk, "nama_atlet": bagan.juara_3b.nama_atlet, "perguruan": bagan.juara_3b.perguruan.nama_perguruan, "juara": "3b", "nama_bagan": bagan.nama_bagan}))
    
    utusans = list(utusans)

    utusans.sort(key=lambda u: (
        -utusan_medals[u.pk]["gold"],
        -utusan_medals[u.pk]["silver"],
        -utusan_medals[u.pk]["bronze"],
    ))

    for utusan in utusans:
        utusan.winners = []

    for winner in utusan_winners:
        for utusan in utusans:
            if utusan.pk == winner['pk']:
                utusan.winners.append(winner)

    context = {
        'on': 'utusan',
        'event': event,
        'utusans': utusans,
        'utusan_medals': utusan_medals,
        'admin_tatami': admin_tatami,
    }
    return render(request, 'admin/utusan.html', context)

def admin_perguruan(request, event_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    perguruans = Perguruan.objects.filter(event=event)
    
    perguruan_medals = defaultdict(lambda: {"gold": 0, "silver": 0, "bronze": 0})
    perguruan_winners = []

    bagans = Bagan.objects.filter(event=event)

    for bagan in bagans:
        if bagan.juara_1 and bagan.juara_1.perguruan:
            perguruan_medals[bagan.juara_1.perguruan.pk]["gold"] += 1
            perguruan_winners.append(({"pk": bagan.juara_1.perguruan.pk, "nama_atlet": bagan.juara_1.nama_atlet, "utusan": bagan.juara_1.utusan.nama_utusan, "juara": "1", "nama_bagan": bagan.nama_bagan}))
        if bagan.juara_2 and bagan.juara_2.perguruan:
            perguruan_medals[bagan.juara_2.perguruan.pk]["silver"] += 1
            perguruan_winners.append(({"pk": bagan.juara_2.perguruan.pk, "nama_atlet": bagan.juara_2.nama_atlet, "utusan": bagan.juara_2.utusan.nama_utusan, "juara": "2", "nama_bagan": bagan.nama_bagan}))
        if bagan.juara_3a and bagan.juara_3a.perguruan:
            perguruan_medals[bagan.juara_3a.perguruan.pk]["bronze"] += 1
            perguruan_winners.append(({"pk": bagan.juara_3a.perguruan.pk, "nama_atlet": bagan.juara_3a.nama_atlet, "utusan": bagan.juara_3a.utusan.nama_utusan, "juara": "3a", "nama_bagan": bagan.nama_bagan}))
        if bagan.juara_3b and bagan.juara_3b.perguruan:
            perguruan_medals[bagan.juara_3b.perguruan.pk]["bronze"] += 1
            perguruan_winners.append(({"pk": bagan.juara_3b.perguruan.pk, "nama_atlet": bagan.juara_3b.nama_atlet, "utusan": bagan.juara_3b.utusan.nama_utusan, "juara": "3b", "nama_bagan": bagan.nama_bagan}))
    
    perguruans = list(perguruans)
    perguruans.sort(key=lambda u: (
        -perguruan_medals[u.pk]["gold"],
        -perguruan_medals[u.pk]["silver"],
        -perguruan_medals[u.pk]["bronze"],
    ))

    for perguruan in perguruans:
        perguruan.winners = []

    for winner in perguruan_winners:
        for perguruan in perguruans:
            if perguruan.pk == winner['pk']:
                perguruan.winners.append(winner)


    context = {
        'on': 'perguruan',
        'event': event,
        'perguruans': perguruans,
        'perguruan_medals': perguruan_medals,
        'admin_tatami': admin_tatami,
    }
    return render(request, 'admin/perguruan.html', context)

def admin_rekapan(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    days = TimetableDay.objects.filter(event=event).order_by('order')

    days_for_filter = [{'pk': d.pk, 'label': format_day_label(d)} for d in days]

    selected_day_ids = request.GET.getlist('day')
    if not selected_day_ids:
        # no filter applied yet (first visit) -> default to showing everything
        selected_day_ids = [str(d.pk) for d in days]
    selected_set = set(selected_day_ids)

    # map: nomor_tanding_id -> set of day_ids it's scheduled on, via the timetable
    nt_day_map = {}
    cells = (
        TimetableCell.objects
        .filter(row__day__event=event, nomor_tanding__isnull=False)
        .select_related('row__day')
    )
    for cell in cells:
        nt_day_map.setdefault(cell.nomor_tanding_id, set()).add(str(cell.row.day_id))

    all_bagans = (
        Bagan.objects.filter(event=event)
        .select_related(
            'juara_1__perguruan', 'juara_1__utusan',
            'juara_2__perguruan', 'juara_2__utusan',
            'juara_3a__perguruan', 'juara_3a__utusan',
            'juara_3b__perguruan', 'juara_3b__utusan',
        ).order_by('kode')
    )

    bagans = []
    for b in all_bagans:
        scheduled_days = nt_day_map.get(b.nomor_tanding_id)
        if not scheduled_days:
            # category isn't placed on the timetable at all yet -> always show,
            # so nothing silently disappears just because scheduling hasn't happened
            bagans.append(b)
        elif scheduled_days & selected_set:
            bagans.append(b)

    context = {
        'event': event,
        'bagans': bagans,
        'days_for_filter': days_for_filter,
        'selected_day_ids': selected_set,
        'admin_tatami': admin_tatami,
    }
    return render(request, 'admin/rekapan.html', context)

def admin_tatami(request, event_pk):
    event = Event.objects.get(pk=event_pk)
    admin_tatami = AdminTatami.objects.filter(user=request.user, event=event).first()
    tatamis = Tatami.objects.filter(event=event)

    if request.method == 'POST':
        if request.POST.get('submit_type') == 'tambah_tatami':
            last_tatami = tatamis.order_by('-tatami_number').first()
            next_number = (last_tatami.tatami_number + 1) if last_tatami else 1

            new_tatami = Tatami.objects.create(event=event, tatami_number=next_number)

            user = User.objects.create_user(username=f'admtatami{next_number}e{event_pk}', password=f'admtatami{next_number}e{event_pk}')
            new_admtatami = AdminTatami.objects.create(event=event, tatami=new_tatami, user=user)

            for i in range(1, 8):
                user = User.objects.create_user(username=f'j{i}t{next_number}e{event_pk}', password=f'j{i}t{next_number}e{event_pk}')
                new_jury = Jury.objects.create(event=event, tatami=new_tatami, user=user, jury_number=i)

            messages.success(request, f"Sukses menambahkan tatami {new_tatami}!")
            return redirect('admin-tatami', event_pk=event_pk)
        elif request.POST.get('submit_type') == 'hapus_tatami':
            tatami = Tatami.objects.filter(pk=request.POST.get('tatami_pk')).first()
            juries = Jury.objects.filter(tatami=tatami)
            adm_tatami = AdminTatami.objects.filter(tatami=tatami).first()
            adm_tatami.user.delete()
            for jury in juries:
                jury.user.delete()
            tatami.delete()
            messages.success(request, f"Sukses menghapus tatami {tatami}!")
            return redirect('admin-tatami', event_pk=event_pk)

    context = {
        'on': 'tatami',
        'event': event,
        'admin_tatami': admin_tatami,
        'tatamis': tatamis,
    }

    return render(request, 'admin/tatami.html', context)

def scoring_board(request, tatami_pk):
    tatami = Tatami.objects.get(pk=tatami_pk)
    
    context = {
        'tatami': tatami,
    }
    return render(request, 'admin/scoring-board.html', context)

@require_POST
def notify_bagan_running(request, detailbagan_pk):
    detail_bagan = DetailBagan.objects.filter(pk=detailbagan_pk).first()
    if not detail_bagan:
        return JsonResponse({'success': False, 'message': 'DetailBagan tidak ditemukan'}, status=404)

    payload = {
        'status': 'running',
        'detail_bagan_id': detail_bagan.pk,
        'bagan_id': detail_bagan.bagan.pk,
        'round': detail_bagan.round,
        'urutan': detail_bagan.urutan,
        'vr1': detail_bagan.vr1,
        'vr2': detail_bagan.vr2,
        'kode_realtime': f'{detail_bagan.bagan.pk}-{detail_bagan.pk}',
        'ring_number': Tatami.objects.filter(detail_bagan=detail_bagan).first().tatami_number,
    }
    if detail_bagan.round != 10:
        success, result = send_to_hosted(payload, endpoint='api/status/')
    return JsonResponse({'success': success, 'message': result})


@require_POST
def send_bagan_result(request, detailbagan_pk):
    detail_bagan = DetailBagan.objects.filter(pk=detailbagan_pk).first()
    if not detail_bagan:
        return JsonResponse({'success': False, 'message': 'DetailBagan tidak ditemukan'}, status=404)

    payload = {
        'status': 'finished',
        'detail_bagan_id': detail_bagan.pk,
        'pemenang': detail_bagan.pemenang,
        'score1': detail_bagan.score1,
        'score2': detail_bagan.score2,
    }
    success, result = send_to_hosted(payload, endpoint='api/result/')
    return JsonResponse({'success': success, 'message': result})

# SORT ------------------------------------------------
AGE_ORDER = [
    'pra usia dini',
    'usia dini',
    'pra pemula',
    'pemula',
    'kadet',
    'junior',
    'senior',
]
AGE_LABELS = {
    'pra usia dini': 'Pra Usia Dini', 'usia dini': 'Usia Dini',
    'pra pemula': 'Pra Pemula', 'pemula': 'Pemula',
    'kadet': 'Kadet', 'junior': 'Junior', 'senior': 'Senior',
}

def get_age_index(name):
    name_lower = name.lower()
    for i, age in enumerate(AGE_ORDER):
        if age in name_lower:
            return i
    return len(AGE_ORDER)
def get_type_index(name):
    name_lower = name.lower()
    is_kata = 'kata' in name_lower
    is_kumite = 'kumite' in name_lower
    is_beregu = 'beregu' in name_lower
    is_putra = 'putra' in name_lower
    is_putri = 'putri' in name_lower
    if is_kata and not is_beregu:
        base = 0 
    elif is_kata and is_beregu:
        base = 2 
    elif is_kumite and not is_beregu:
        base = 4 
    elif is_kumite and is_beregu:
        base = 6 
    else:
        base = 8 

    if is_putra:
        return base
    elif is_putri:
        return base + 1
    return base + 0.5

def get_weight_key(name):
    match = re.search(r'([+-])\s*(\d+)\s*kg', name.lower())
    if match:
        sign, num = match.groups()
        sign_rank = 0 if sign == '-' else 1
        return (sign_rank, int(num))
    return (0.5, 0)  # no weight class in the name (e.g. Kata categories) — neutral, doesn't disturb ordering

def sort_key(nomor_tanding):
    name = nomor_tanding.nama_nomor_tanding or ''
    sign_rank, weight_num = get_weight_key(name)
    return (get_age_index(name), get_type_index(name), sign_rank, weight_num, name)

def create_bagan_and_seed(event, nomor_tanding, nama_bagan, pool, group_counts, atlets_temp, group_field, custom_order, atlet_assignment1):
    """Stage 1 + 2: create the Bagan, then assign athletes into round 1."""
    bagan = Bagan.objects.create(
        event=event, nomor_tanding=nomor_tanding, nama_bagan=nama_bagan, pool=pool
    )
    name = nomor_tanding.nama_nomor_tanding or ''
    if 'KATA' in name:
        bagan.tipe_tanding = '1'
    elif 'KUMITE' in name:
        bagan.tipe_tanding = '2'
    bagan.save()

    _seed_round_1(bagan, custom_order, group_counts, atlets_temp, group_field, atlet_assignment1)
    return bagan


def _pick_and_place(detail_bagan, group_counts, atlets_temp, group_field, group_index, target_field):
    """Shared inner loop: find an eligible athlete for a group, assign it, advance group_index."""
    assigned = False
    while group_index < len(group_counts) and not assigned:
        group_id, remaining = group_counts[group_index]
        eligible = [a for a in atlets_temp if getattr(a, group_field + '_id') == group_id]

        if eligible:
            atlet = random.choice(eligible)
            setattr(detail_bagan, target_field, atlet)
            setattr(detail_bagan, 'vr1' if target_field == 'atlet1' else 'vr2', True)
            detail_bagan.save()

            atlets_temp.remove(atlet)
            group_counts[group_index] = (group_id, remaining - 1)
            if group_counts[group_index][1] <= 0:
                group_index += 1
            assigned = True
        else:
            group_index += 1
    return group_index


def _seed_round_1(bagan, custom_order, group_counts, atlets_temp, group_field, atlet_assignment1):
    # Pass 1: create every round-1 slot, filling in athletes as we go
    group_index = 0
    for urutan in custom_order:
        detail_bagan = DetailBagan.objects.create(bagan=bagan, round=1, urutan=urutan)
        if atlets_temp and group_index < len(group_counts):
            target_field = atlet_assignment1.get(urutan)
            group_index = _pick_and_place(detail_bagan, group_counts, atlets_temp, group_field, group_index, target_field)

    # Pass 2: backfill any slot left with only one athlete assigned
    group_index = 0
    for urutan in custom_order:
        if not atlets_temp:
            break
        detail_bagan = DetailBagan.objects.get(bagan=bagan, round=1, urutan=urutan)
        target_field = atlet_assignment1.get(urutan)

        if target_field == 'atlet1' and detail_bagan.atlet2 is None:
            eligible_field = 'atlet2'
        elif target_field == 'atlet2' and detail_bagan.atlet1 is None:
            eligible_field = 'atlet1'
        else:
            continue

        group_index = _pick_and_place(detail_bagan, group_counts, atlets_temp, group_field, group_index, eligible_field)


def _advance_round_1_to_2(bagan):
    urutan_map = {1: (1, 'atlet1'), 2: (1, 'atlet2'), 3: (2, 'atlet1'), 4: (2, 'atlet2'),
                  5: (3, 'atlet1'), 6: (3, 'atlet2'), 7: (4, 'atlet1'), 8: (4, 'atlet2')}

    for detail_bagan in DetailBagan.objects.filter(bagan=bagan, round=1).order_by('urutan'):
        target = urutan_map.get(detail_bagan.urutan)
        if not target:
            continue
        target_urutan, target_field = target

        new_detail_bagan, _ = DetailBagan.objects.get_or_create(bagan=bagan, round=2, urutan=target_urutan)

        atlet = None
        if not detail_bagan.atlet1:
            atlet, detail_bagan.atlet2, detail_bagan.vr2 = detail_bagan.atlet2, None, False
        elif not detail_bagan.atlet2:
            atlet, detail_bagan.atlet1, detail_bagan.vr1 = detail_bagan.atlet1, None, False

        if atlet:
            setattr(new_detail_bagan, target_field, atlet)
            setattr(new_detail_bagan, 'vr1' if target_field == 'atlet1' else 'vr2', True)

        detail_bagan.save()
        new_detail_bagan.save()


def _advance_round_2_to_3(bagan):
    match_map = {
        1: {'round1_urutans': (2, 1), 'round3_urutan': 1, 'slot': 'atlet1'},
        2: {'round1_urutans': (4, 3), 'round3_urutan': 1, 'slot': 'atlet2'},
        3: {'round1_urutans': (6, 5), 'round3_urutan': 2, 'slot': 'atlet1'},
        4: {'round1_urutans': (8, 7), 'round3_urutan': 2, 'slot': 'atlet2'},
    }

    for detail_bagan in DetailBagan.objects.filter(bagan=bagan, round=2).order_by('urutan'):
        config = match_map.get(detail_bagan.urutan)
        if not config:
            continue

        round1_a, round1_b = config['round1_urutans']
        target_slot = config['slot']
        new_detail_bagan, _ = DetailBagan.objects.get_or_create(bagan=bagan, round=3, urutan=config['round3_urutan'])
        if bagan.pool == 1:
            new_detail_bagan.vr1 = True
            new_detail_bagan.vr2 = True

        if not detail_bagan.atlet2:
            target_detail = DetailBagan.objects.filter(bagan=bagan, round=1, urutan=round1_a).first()
            if target_detail and (not target_detail.atlet1 or not target_detail.atlet2):
                setattr(new_detail_bagan, target_slot, detail_bagan.atlet1)
                detail_bagan.atlet1, detail_bagan.vr1 = None, False
                setattr(new_detail_bagan, 'vr1' if target_slot == 'atlet1' else 'vr2', True)

        elif not detail_bagan.atlet1:
            target_detail = DetailBagan.objects.filter(bagan=bagan, round=1, urutan=round1_b).first()
            if target_detail and (not target_detail.atlet1 or not target_detail.atlet2):
                setattr(new_detail_bagan, target_slot, detail_bagan.atlet2)
                detail_bagan.atlet2, detail_bagan.vr2 = None, False
                setattr(new_detail_bagan, 'vr1' if target_slot == 'atlet1' else 'vr2', True)

        detail_bagan.save()
        new_detail_bagan.save()


def _advance_round_3_to_4(bagan):
    # config per round-3 urutan: which round-2 match determines "still-contested",
    # which slot on the round-4 final we write to, and which side's athlete/vr to check first
    config_map = {
        1: {'check_urutans': (2, 1), 'slot_if_atlet2_missing': 'atlet1', 'slot_if_atlet1_missing': 'atlet1'},
        2: {'check_urutans': (4, 3), 'slot_if_atlet2_missing': 'atlet2', 'slot_if_atlet1_missing': 'atlet2'},
    }

    for detail_bagan in DetailBagan.objects.filter(bagan=bagan, round=3).order_by('urutan'):
        config = config_map.get(detail_bagan.urutan)
        if not config:
            continue

        new_detail_bagan = DetailBagan.objects.filter(bagan=bagan, round=4, urutan=1).first()
        if not new_detail_bagan:
            new_detail_bagan = DetailBagan.objects.create(bagan=bagan, round=4, urutan=1)
            if bagan.pool in (1, 2):
                new_detail_bagan.vr1 = True
                new_detail_bagan.vr2 = True

        check_a, check_b = config['check_urutans']

        if not detail_bagan.atlet2:
            detail_bagan.vr2 = False
            check_match = DetailBagan.objects.filter(bagan=bagan, round=2, urutan=check_a).first()
            if check_match and (not check_match.atlet1 or not check_match.atlet2):
                slot = config['slot_if_atlet2_missing']
                setattr(new_detail_bagan, slot, detail_bagan.atlet1)
                setattr(new_detail_bagan, 'vr1' if slot == 'atlet1' else 'vr2', True)
                detail_bagan.atlet1, detail_bagan.vr1 = None, False

        elif not detail_bagan.atlet1:
            detail_bagan.vr1 = False
            check_match = DetailBagan.objects.filter(bagan=bagan, round=2, urutan=check_b).first()
            if check_match and (not check_match.atlet1 or not check_match.atlet2):
                slot = config['slot_if_atlet1_missing']
                setattr(new_detail_bagan, slot, detail_bagan.atlet2)
                setattr(new_detail_bagan, 'vr1' if slot == 'atlet1' else 'vr2', True)
                detail_bagan.atlet2, detail_bagan.vr2 = None, False

        detail_bagan.save()
        new_detail_bagan.save()

def build_full_bracket(event, nomor_tanding, nama_bagan, pool, group_counts, atlets_temp, group_field, custom_order, atlet_assignment1):
    """Runs the entire pipeline for one Bagan (one pool or the whole category) and returns (bagan, round5_detail_bagan)."""
    bagan = create_bagan_and_seed(event, nomor_tanding, nama_bagan, pool, group_counts, atlets_temp, group_field, custom_order, atlet_assignment1)
    _advance_round_1_to_2(bagan)
    _advance_round_2_to_3(bagan)
    _advance_round_3_to_4(bagan)
    round_5 = DetailBagan.objects.create(bagan=bagan, round=5, urutan=1)
    return bagan, round_5

def get_age_group(name):
    name_lower = (name or '').lower()
    for age in AGE_ORDER:
        if age in name_lower:
            return AGE_LABELS[age]
    return 'Lainnya'

INDO_MONTHS = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 'Juli',
               'Agustus', 'September', 'Oktober', 'November', 'Desember']

def format_day_label(day):
    day_number = day.order + 1
    if day.tanggal:
        t = day.tanggal
        return f"Day {day_number}, {t.day} {INDO_MONTHS[t.month - 1]} {t.year}"
    return f"Day {day_number}"

def timetable_editor(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    tatamis = Tatami.objects.filter(event=event).order_by('tatami_number')

    days = TimetableDay.objects.filter(event=event).prefetch_related('rows__cells')

    if not days.exists():
        first_day = TimetableDay.objects.create(event=event, order=0)
        days = TimetableDay.objects.filter(event=event).prefetch_related('rows__cells')

    atlet_counts = dict(
        Atlet.objects.filter(nomor_tanding__event=event)
        .values('nomor_tanding').annotate(cnt=Count('id'))
        .values_list('nomor_tanding', 'cnt')
    )
    nomor_tanding_qs = NomorTanding.objects.filter(event=event).order_by('nama_nomor_tanding')
    nt_with_group = [
        {'pk': nt.pk, 'nama': nt.nama_nomor_tanding, 'age_group': get_age_group(nt.nama_nomor_tanding),
         'atlet_count': atlet_counts.get(nt.pk, 0)}
        for nt in nomor_tanding_qs
    ]
    ordered_labels = [AGE_LABELS[a] for a in AGE_ORDER] + ['Lainnya']
    age_groups = [g for g in ordered_labels if any(nt['age_group'] == g for nt in nt_with_group)]

    days_data = []
    for day in days:
        cell_map = {row.id: {c.tatami_id: c for c in row.cells.all()} for row in day.rows.all()}
        days_data.append({
            'pk': day.pk,
            'label': format_day_label(day),
            'tanggal': day.tanggal.isoformat() if day.tanggal else '',
            'rows': day.rows.all(),
            'cell_map': cell_map,
        })

    kop, _ = KopSurat.objects.get_or_create(event=event)
    kop_surat_data = {
        'logo_url': kop.logo.url if kop.logo else '',
        'nama_organisasi': kop.nama_organisasi,
        'alamat': kop.alamat,
        'kontak': kop.kontak,
    }

    keterangan, _ = EventKeterangan.objects.get_or_create(event=event)

    context = {
        'event': event, 'on': 'roster-maker', 'tatamis': tatamis,
        'days_data': days_data, 'nt_with_group': nt_with_group, 'age_groups': age_groups,
        'kop_surat': kop_surat_data,
        'atlet_counts': atlet_counts, 'keterangan_text': keterangan.text
    }
    return render(request, 'admin/timetable_editor.html', context)


@require_POST
def add_day(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    next_order = TimetableDay.objects.filter(event=event).count()
    day = TimetableDay.objects.create(event=event, order=next_order)
    return JsonResponse({'success': True, 'id': day.pk, 'order': day.order, 'label': format_day_label(day)})


@require_POST
def delete_day(request, event_pk, day_pk):
    day = get_object_or_404(TimetableDay, pk=day_pk, event_id=event_pk)
    day.delete()
    return JsonResponse({'success': True})


@require_POST
def timetable_save(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)

    with transaction.atomic():
        for day_data in data.get('days', []):
            day = TimetableDay.objects.filter(pk=day_data.get('day_pk'), event=event).first()
            if not day:
                continue

            tanggal = day_data.get('tanggal') or None
            day.tanggal = tanggal
            day.save()

            TimetableRow.objects.filter(day=day).delete()

            for order, row_data in enumerate(day_data.get('rows', [])):
                row = TimetableRow.objects.create(
                    day=day, order=order,
                    row_type=row_data.get('row_type', 'slot'),
                    time_label=row_data.get('time_label', ''),
                    label_text=row_data.get('label_text', ''),
                )
                if row.row_type == 'slot':
                    for cell_data in row_data.get('cells', []):
                        tatami_id = cell_data.get('tatami_id')
                        if not tatami_id:
                            continue
                        TimetableCell.objects.create(
                            row=row, tatami_id=tatami_id,
                            nomor_tanding_id=cell_data.get('nomor_tanding_id') or None,
                            custom_text=cell_data.get('custom_text', ''),
                        )

    return JsonResponse({'success': True})

@require_POST
def add_tatami(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)

    tatami_number = data.get('tatami_number')
    if not tatami_number:
        return JsonResponse({'success': False, 'message': 'tatami_number is required'}, status=400)

    tatami = Tatami.objects.create(event=event, tatami_number=tatami_number)
    return JsonResponse({'success': True, 'id': tatami.pk, 'tatami_number': tatami.tatami_number})

@require_POST
def delete_tatami(request, event_pk, tatami_pk):
    tatami = get_object_or_404(Tatami, pk=tatami_pk, event_id=event_pk)
    tatami.delete()  # cascades to TimetableCell rows referencing it
    return JsonResponse({'success': True})

def kop_surat_get(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    kop, _ = KopSurat.objects.get_or_create(event=event)
    return JsonResponse({
        'success': True,
        'logo_url': kop.logo.url if kop.logo else '',
        'nama_organisasi': kop.nama_organisasi,
        'alamat': kop.alamat,
        'kontak': kop.kontak,
    })

@require_POST
def kop_surat_save(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    kop, _ = KopSurat.objects.get_or_create(event=event)

    kop.nama_organisasi = request.POST.get('nama_organisasi', '')
    kop.alamat = request.POST.get('alamat', '')
    kop.kontak = request.POST.get('kontak', '')

    if request.FILES.get('logo'):
        kop.logo = request.FILES['logo']

    kop.save()
    return JsonResponse({
        'success': True,
        'logo_url': kop.logo.url if kop.logo else '',
    })

@require_POST
def keterangan_save(request, event_pk):
    event = get_object_or_404(Event, pk=event_pk)
    keterangan, _ = EventKeterangan.objects.get_or_create(event=event)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)

    keterangan.text = data.get('text', '')
    keterangan.save()
    return JsonResponse({'success': True})

def get_ordered_bagans_for_tatami(day, tatami):
    rows = (
        TimetableRow.objects
        .filter(day=day, row_type='slot')
        .order_by('order')
        .prefetch_related('cells')
    )

    seen_nt_ids = []
    for row in rows:
        cell = row.cells.filter(tatami=tatami).first()
        if cell and cell.nomor_tanding_id and cell.nomor_tanding_id not in seen_nt_ids:
            seen_nt_ids.append(cell.nomor_tanding_id)

    ordered_bagans = []
    for nt_id in seen_nt_ids:
        # a single nomor_tanding can have multiple Bagan objects (Pool A/B/C/D + Final)
        bagans = list(Bagan.objects.filter(nomor_tanding_id=nt_id, event=day.event))
        bagans.sort(key=lambda b: (1 if b.pool == 0 else 0, b.nama_bagan or ''))
        ordered_bagans.extend(bagans)

    return ordered_bagans

def _render_pdf_worker(session_cookie_name, session_cookie_value, base_url, target_url):
    if sys.platform.startswith('win'):
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

    with sync_playwright() as p:
        browser = p.chromium.launch()
        context = browser.new_context()

        if session_cookie_value:
            context.add_cookies([{
                'name': session_cookie_name,
                'value': session_cookie_value,
                'url': base_url,
            }])

        page = context.new_page()
        page.goto(target_url, wait_until='networkidle')
        pdf_bytes = page.pdf(
            format='A4',
            landscape=True,
            print_background=True,
            margin={'top': '0', 'bottom': '0', 'left': '0', 'right': '0'},
        )
        browser.close()
        return pdf_bytes


def render_authenticated_page_to_pdf(request, url):
    session_cookie_value = request.COOKIES.get(settings.SESSION_COOKIE_NAME)
    base_url = request.build_absolute_uri('/')

    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(
            _render_pdf_worker,
            settings.SESSION_COOKIE_NAME,
            session_cookie_value,
            base_url,
            url,
        )
        return future.result()
    
def bulk_print_bagan(request, event_pk, day_pk, tatami_pk):
    event = get_object_or_404(Event, pk=event_pk)
    day = get_object_or_404(TimetableDay, pk=day_pk, event=event)
    tatami = get_object_or_404(Tatami, pk=tatami_pk, event=event)

    bagans = get_ordered_bagans_for_tatami(day, tatami)
    if not bagans:
        return HttpResponse('Tidak ada bagan terjadwal di tatami ini.', status=404)

    writer = PdfWriter()
    for bagan in bagans:
        url = request.build_absolute_uri(
            reverse('admin-bagan-detail', kwargs={'event_pk': event.pk, 'bagan_pk': bagan.pk}) + '?print=1'
        )
        pdf_bytes = render_authenticated_page_to_pdf(request, url)
        reader = PdfReader(io.BytesIO(pdf_bytes))
        for page in reader.pages:
            writer.add_page(page)

    buffer = io.BytesIO()
    writer.write(buffer)
    buffer.seek(0)

    filename = f"Day{day.order + 1}_Tatami_{tatami.tatami_number}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response